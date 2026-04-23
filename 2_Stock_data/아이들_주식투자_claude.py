import yfinance as yf
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

EXCEL_PATH = r"C:\Users\MSJ\Desktop\아이들_주식투자_claude_cursor.xlsx"
SHEET_NAME = "stock_price"

TICKERS = ["GOOGL", "MSFT", "NVDA", "QQQ", "XLK", "AAPL", "GOOGL", "XLK"]
UNIQUE_TICKERS = list(dict.fromkeys(TICKERS))  # 중복 제거 (순서 유지)
USDKRW_TICKER = "USDKRW=X"  # Yahoo Finance: USD 1달러당 KRW 가격

def fetch_usdkrw_rate():
    try:
        fx_stock = yf.Ticker(USDKRW_TICKER)
        hist = fx_stock.history(period="5d")
        if len(hist) >= 1:
            return float(hist["Close"].iloc[-1])
    except Exception as e:
        print(f"환율 데이터 오류: {e}")
    return None

def fetch_stock_data(tickers):
    results = {}
    for ticker in tickers:
        try:
            stock = yf.Ticker(ticker)
            hist = stock.history(period="5d")
            if len(hist) >= 2:
                current_price = hist["Close"].iloc[-1]
                prev_price = hist["Close"].iloc[-2]
                change_pct = (current_price - prev_price) / prev_price * 100
            elif len(hist) == 1:
                current_price = hist["Close"].iloc[-1]
                change_pct = None
            else:
                current_price = None
                change_pct = None

            # 배당률(티커별 dividendYield/trailingAnnualDividendYield)
            dividend_yield_ratio = None  # 엑셀 표시용: 0.0123 형태(=1.23%)
            try:
                info = stock.info or {}
                dy = info.get("dividendYield", None)
                if dy is None:
                    dy = info.get("trailingAnnualDividendYield", None)

                if dy is not None:
                    # 안전하게 정규화 후, 요청: 값이 100배 높게 나오는 케이스 대비
                    dividend_yield_ratio = (dy / 100 if dy > 1 else dy) / 100
            except Exception:
                dividend_yield_ratio = None

            results[ticker] = {
                "price": current_price,
                "change_pct": change_pct,
                "dividend_yield_ratio": dividend_yield_ratio,
            }
        except Exception as e:
            print(f"{ticker} 데이터 오류: {e}")
            results[ticker] = {"price": None, "change_pct": None, "dividend_yield_ratio": None}
    return results

def apply_header_style(cell, bg_color="1F4E79"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_border(cell):
    thin = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def get_change_color(change_pct):
    if change_pct is None:
        return None
    if change_pct > 0:
        return "C6EFCE"  # 연두색 (상승)
    elif change_pct < 0:
        return "FFC7CE"  # 연빨간색 (하락)
    return None

def write_to_excel(stock_data, tickers):
    # 원본 엑셀 파일이 없으면 새 워크북으로 시작
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()

    fx_rate = fetch_usdkrw_rate()

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    # 헤더 작성
    headers = ["종목", "현재가 (USD)", "변동률 (%)", "배당률 (%)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        apply_border(cell)

    # 업데이트 날짜/시간 표기 (요청: G1=날짜, H1=시간)
    now = datetime.now()
    cell_date = ws.cell(row=1, column=7, value=now.strftime("%Y-%m-%d"))
    cell_time = ws.cell(row=1, column=8, value=now.strftime("%H:%M:%S"))
    for c in (cell_date, cell_time):
        c.font = Font(italic=True, color="808080", name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(c)

    # 업데이트 헤더 옆 환율 라벨/값 (요청: G2=USD/KRW, H2=원달러 환율)
    cell_fx_label = ws.cell(row=2, column=7, value="USD/KRW")
    cell_fx_value = ws.cell(row=2, column=8, value=fx_rate if fx_rate is not None else "N/A")
    for c in (cell_fx_label, cell_fx_value):
        c.font = Font(italic=True, color="808080", name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(c)
    if fx_rate is not None:
        cell_fx_value.number_format = '#,##0.00'

    # 데이터 작성 (요청한 순서대로, 중복 포함)
    for row, ticker in enumerate(tickers, start=2):
        data = stock_data.get(ticker, {})
        price = data.get("price")
        change_pct = data.get("change_pct")
        dividend_yield_ratio = data.get("dividend_yield_ratio")

        # 종목명
        cell_ticker = ws.cell(row=row, column=1, value=ticker)
        cell_ticker.font = Font(bold=True, name="Arial", size=11)
        cell_ticker.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(cell_ticker)

        # 현재가
        cell_price = ws.cell(row=row, column=2, value=round(price, 2) if price else "N/A")
        cell_price.font = Font(name="Arial", size=11)
        cell_price.alignment = Alignment(horizontal="right", vertical="center")
        if price:
            cell_price.number_format = '#,##0.00'
        apply_border(cell_price)

        # 변동률
        if change_pct is not None:
            cell_change = ws.cell(row=row, column=3, value=round(change_pct, 2))
            cell_change.number_format = '+0.00%;-0.00%;0.00%'
            cell_change.value = change_pct / 100  # 퍼센트 형식으로 저장
            bg = get_change_color(change_pct)
            if bg:
                cell_change.fill = PatternFill("solid", start_color=bg)
        else:
            cell_change = ws.cell(row=row, column=3, value="N/A")
        cell_change.font = Font(name="Arial", size=11)
        cell_change.alignment = Alignment(horizontal="right", vertical="center")
        apply_border(cell_change)

        # 배당률
        if dividend_yield_ratio is not None:
            cell_div = ws.cell(row=row, column=4, value=dividend_yield_ratio)
            cell_div.number_format = '0.00%'
        else:
            cell_div = ws.cell(row=row, column=4, value="N/A")
        cell_div.font = Font(name="Arial", size=11)
        cell_div.alignment = Alignment(horizontal="right", vertical="center")
        apply_border(cell_div)

    # 열 너비 조정
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12

    # 행 높이 조정
    ws.row_dimensions[1].height = 22
    for row in range(2, len(tickers) + 2):
        ws.row_dimensions[row].height = 20

    try:
        wb.save(EXCEL_PATH)
        saved_path = EXCEL_PATH
    except PermissionError:
        # 엑셀이 열려있어 원본 파일 쓰기 권한이 막히는 경우 대비
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        saved_path = EXCEL_PATH.replace(".xlsx", f"_{ts}.xlsx")
        wb.save(saved_path)
    print(f"'{SHEET_NAME}' 시트가 '{saved_path}'에 저장되었습니다.")

if __name__ == "__main__":
    print("주가 데이터를 가져오는 중...")
    stock_data = fetch_stock_data(UNIQUE_TICKERS)

    print("\n[현재가 조회 결과]")
    for ticker, data in stock_data.items():
        price = data['price']
        chg = data['change_pct']
        price_str = f"${price:.2f}" if price else "N/A"
        chg_str = f"{chg:+.2f}%" if chg is not None else "N/A"
        print(f"  {ticker:6s} | {price_str:>10} | {chg_str:>8}")

    write_to_excel(stock_data, TICKERS)