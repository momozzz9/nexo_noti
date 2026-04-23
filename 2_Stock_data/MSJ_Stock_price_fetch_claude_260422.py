import sys
import time
import requests
import yfinance as yf
import pandas as pd
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Alignment, PatternFill, Font
from pycoingecko import CoinGeckoAPI

sys.stdout.reconfigure(encoding="utf-8")

script_start = time.time()

# ----------------------
# 1. 종목 리스트
# ----------------------
kr_stocks = [
    ("Time 나스닥 채권혼합",          "0019K0"),
    ("Time 차이나 AI테크",            "0043Y0"),
    ("Plus 금 채권혼합",              "0138Y0"),
    ("Kodex 미국반도체",              "390390"),
    ("Tiger KRX 금현물",             "411060"),
    ("Time 미국나스닥100",            "426030"),
    ("Kodex 미국배당 커버드콜",        "441640"),
    ("Time Korea 플러스 배당",         "441800"),
    ("Time 글로벌 AI 인공지능",        "456600"),

    ("Plus K 방산",                  "449450"),
    ("Sol 조선 Top3 플러스",          "466920"),
    ("Tiger 반도체 Top10",            "396500"),

    ("Kodex 미국반도체",              "390390"),
    ("Plus 글로벌 HBM 반도체",        "442580"),
    ("Rise 미국AI밸류체인 커버드콜",    "490590"),
    ("Time 글로벌 AI 인공지능",        "456600"),
    ("Time 글로벌 우주테크&방산",      "478150"),
    ("Time 미국나스닥100",            "426030"),
    ("Time 차이나 AI테크",            "0043Y0"),

    ("Time 차이나 AI테크",            "0043Y0"),
    ("Kodex 미국반도체",              "390390"),
    ("Time 미국나스닥100",            "426030"),
    ("Time Korea 플러스 배당",         "441800"),
    ("Time 글로벌 AI 인공지능",        "456600"),

    ("Time 나스닥 채권혼합",          "0019K0"),
    ("Time 미국나스닥100",            "426030"),

    ("Kiwoom 미국원유에너지 기업",     "474800"),
    ("Kodex 200 타겟위클리 커버드콜",  "498400"),
    ("Kodex AI전력핵심설비",          "487240"),
    ("Kodex 미국우주항공",            "0167Z0"),
    ("Kodex 차이나 AI반도체",         "0162L0"),
    ("Rise 미국AI밸류체인 커버드콜",    "490590"),
    ("Tiger반도체 Top10 레버리지",    "488080"),
    ("Tiger 차이나 과창판50",         "414780"),
    ("Tiger 항셍테크",               "371160"),
]

us_stocks = [
    "GDE", "ITA", "QLD", "QQQM", "QTUM", "SCHD", "SMH", "XLE",
    "QDVO", "QQQI", "SCHD",
    "AIPO", "GDX", "IDEF", "KORU", "SLV", "SOXL", "TQQQ", "ULTY",
]

jp_stocks = [
    ("일본 종합상사", "1629.T"),
    ("일본 보험사",  "1632.T"),
    ("일본 반도체",  "200A.T"),
]

kr_groups = [
    ("<삼성증권 퇴직연금>", kr_stocks[0:9]),
    ("<Toss>",            kr_stocks[9:12]),
    ("<NH농협 개인연금>",  kr_stocks[12:19]),
    ("<미래에셋 개인연금>", kr_stocks[19:24]),
    ("<미래에셋 IRP>",    kr_stocks[24:26]),
    ("<NH ISA>",         kr_stocks[26:]), # 마지막까지 포함하도록 수정
]

us_groups = [
    ("<Creon>",  us_stocks[0:8]),
    ("<Meritz>", us_stocks[8:11]),
    ("<Toss>",   us_stocks[11:]), # 마지막까지 포함하도록 수정
]

# ----------------------
# 2. 데이터 fetch 함수 (종목 1개 → 가격 + 변동률 + 배당률)
# ----------------------
def fetch_stock(ticker_str: str) -> dict:
    """
    yfinance를 사용하여 단일 종목의 데이터를 가져옵니다.
    가격, 전일 대비 변동률, 배당률을 추출하며, 데이터가 불완전할 경우 history 데이터를 통해 보완합니다.
    한국 주식 및 일부 ETF의 경우 배당 이력을 직접 합산하여 배당률을 계산합니다.
    """
    # 기본값 설정
    result = {"close": None, "change_pct": "-", "dividend_yield": None}
    
    try:
        t = yf.Ticker(ticker_str)
        # info API는 실시간 데이터나 일부 필드가 누락되는 경우가 많음
        info = t.info
        
        # 1. 현재가 추출 (여러 키값을 순차적으로 시도)
        close = (info.get("regularMarketPrice") or 
                 info.get("currentPrice") or 
                 info.get("navPrice") or 
                 info.get("previousClose")) # 마지막 수단으로 전일종가라도 시도
        
        # 2. 전일 종가 추출 (변동률 계산용)
        prev = info.get("regularMarketPreviousClose") or info.get("previousClose")
        
        # 3. 데이터 보완 (가격이 없거나 전일 종가가 0 또는 None인 경우 history 사용)
        if close is None or prev is None or prev <= 0:
            hist = t.history(period="7d", interval="1d", timeout=10)
            
            if not hist.empty and len(hist) >= 1:
                if close is None or (isinstance(close, (int, float)) and close <= 0):
                    close = float(hist["Close"].iloc[-1])
                
                if (prev is None or prev <= 0) and len(hist) >= 2:
                    prev = float(hist["Close"].iloc[-2])
            
        # 결과 딕셔너리에 값 반영
        if close is not None and close > 0:
            result["close"] = close
            
            # 4. 변동률 계산
            if prev is not None and prev > 0:
                change = (close - prev) / prev * 100
                result["change_pct"] = f"{change:.2f}%"
            else:
                result["change_pct"] = "-"

        # 5. 배당률 추출 및 계산 (로직 정밀화)
        div_yield_val = None
        ref_yield = None # 비교용 참조 배당률
        
        # (A) 참조값 계산 (배당금액 또는 이력 기반)
        div_rate = info.get("dividendRate")
        if div_rate and close and close > 0:
            ref_yield = (float(div_rate) / float(close)) * 100
        else:
            # 배당금 정보가 없으면 이력에서 계산 시도
            try:
                divs = t.dividends
                if not divs.empty:
                    now = pd.Timestamp.now(tz=divs.index.tz) if divs.index.tz else pd.Timestamp.now()
                    one_year_ago = now - pd.Timedelta(days=365)
                    # Series/DataFrame sum 결과에서 스칼라 값을 추출하기 위해 float() 사용
                    div_sum = divs[divs.index > one_year_ago].sum()
                    last_year_divs = float(div_sum.iloc[0]) if hasattr(div_sum, "iloc") else float(div_sum)
                    
                    if last_year_divs > 0 and close and close > 0:
                        ref_yield = (last_year_divs / close) * 100
            except Exception:
                pass

        # (B) info의 dividendYield 확인 및 교정
        raw_yield = info.get("dividendYield")
        if raw_yield is not None and raw_yield > 0:
            if ref_yield is not None and ref_yield > 0:
                # 참조값이 있는 경우, raw_yield와 raw_yield*100 중 참조값에 더 가까운 쪽 선택
                # 예: ref=0.15, raw=0.19 -> 0.19 선택 / ref=5.0, raw=0.05 -> 5.0 선택
                if abs(raw_yield - ref_yield) < abs(raw_yield * 100 - ref_yield):
                    div_yield_val = raw_yield
                else:
                    div_yield_val = raw_yield * 100
            else:
                # 참조값이 없는 경우의 휴리스틱
                # 최근 yfinance는 퍼센트 단위(3.44 등)를 선호하므로, 
                # 아주 작은 값(0.01 미만)이 아니면 그대로 사용
                if raw_yield < 0.01: 
                    div_yield_val = raw_yield * 100
                else:
                    div_yield_val = raw_yield
        
        # (C) 최종값 결정: info 정보가 없으면 참조값 사용
        if div_yield_val is None or div_yield_val == 0:
            div_yield_val = ref_yield

        # 최종 포맷팅
        if div_yield_val is not None:
            # 비정상적으로 높은 값(예: 500% 초과)은 데이터 오류일 가능성이 높음
            if div_yield_val > 500:
                result["dividend_yield"] = f"{div_yield_val:.2f}% (!)"
            else:
                # 소수점 2째자리까지 표시
                result["dividend_yield"] = f"{div_yield_val:.2f}%"
        else:
            # 데이터가 전혀 없는 경우에도 0.00%로 표시
            result["dividend_yield"] = "0.00%"
                
    except Exception:
        pass
        
    return result


def fetch_all_parallel(ticker_list: list[str], max_workers: int = 20) -> dict[str, dict]:
    """ticker_list를 병렬로 fetch하고 {ticker: result} dict 반환."""
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {ex.submit(fetch_stock, t): t for t in ticker_list}
        for future in as_completed(future_map):
            ticker = future_map[future]
            try:
                results[ticker] = future.result()
            except Exception:
                results[ticker] = {"close": None, "change_pct": "-", "dividend_yield": None}
    return results


# ----------------------
# 3. 고유 ticker 목록 수집 후 한 번에 병렬 fetch
# ----------------------
unique_kr = list({f"{code}.KS" for _, code in kr_stocks})
unique_us = list(set(us_stocks))
unique_jp = list({ticker for _, ticker in jp_stocks})
# 모든 고유 티커를 합쳐서 병렬로 데이터 수집
all_unique = unique_kr + unique_us + unique_jp

print(f"[{time.time()-script_start:.1f}s] 병렬 데이터 수집 시작 ({len(all_unique)}개 고유 종목)...")
# ThreadPoolExecutor를 사용하여 병렬로 데이터를 수집함으로써 전체 실행 시간을 단축합니다.
cache = fetch_all_parallel(all_unique)
print(f"[{time.time()-script_start:.1f}s] 데이터 수집 완료")


# ----------------------
# 4. 외부 API (환율 + 암호화폐) 병렬 fetch
# ----------------------
def fetch_fx():
    """yfinance로 USD 기준 환율 가져오기 (KRW, JPY, CNY, HKD) - API key 불필요"""
    fx_tickers = {
        "KRW": "KRW=X",
        "JPY": "JPY=X",
        "CNY": "CNY=X",
        "HKD": "HKD=X",
    }
    rates = {}
    try:
        tickers_str = " ".join(fx_tickers.values())
        data = yf.download(
            tickers_str,
            period="2d",
            interval="1d",
            group_by="ticker",
            progress=False,
            timeout=10,
        )
        for currency, ticker in fx_tickers.items():
            try:
                price = float(data[ticker]["Close"].dropna().iloc[-1])
                rates[currency] = price
            except Exception:
                rates[currency] = None
    except Exception:
        pass

    # yf.download 실패 시 개별 Ticker로 폴백
    for currency, ticker in fx_tickers.items():
        if rates.get(currency) is None:
            try:
                t = yf.Ticker(ticker)
                info = t.info
                price = info.get("regularMarketPrice") or info.get("regularMarketPreviousClose")
                if price:
                    rates[currency] = float(price)
                else:
                    hist = t.history(period="2d", interval="1d", timeout=5)
                    if not hist.empty:
                        rates[currency] = float(hist["Close"].dropna().iloc[-1])
            except Exception:
                rates[currency] = None

    return rates


def fetch_crypto():
    try:
        cg = CoinGeckoAPI()
        data = cg.get_price(ids="bitcoin,ethereum,tether,nexo", vs_currencies="usd")
        return {
            "BTC":  data.get("bitcoin",  {}).get("usd"),
            "ETH":  data.get("ethereum", {}).get("usd"),
            "USDT": data.get("tether",   {}).get("usd"),
            "NEXO": data.get("nexo",     {}).get("usd"),
        }
    except Exception:
        return {}

def fetch_upbit():
    symbols = ["KRW-BTC", "KRW-ETH", "KRW-USDT"]
    try:
        r = requests.get("https://api.upbit.com/v1/ticker",
                         params={"markets": ",".join(symbols)}, timeout=5)
        r.raise_for_status()
        return {item["market"]: item["trade_price"] for item in r.json()}
    except Exception:
        return {}

print(f"[{time.time()-script_start:.1f}s] 외부 API 병렬 호출...")
with ThreadPoolExecutor(max_workers=3) as ex:
    f_fx     = ex.submit(fetch_fx)
    f_crypto = ex.submit(fetch_crypto)
    f_upbit  = ex.submit(fetch_upbit)
    rate_data    = f_fx.result()
    crypto_data  = f_crypto.result()
    upbit_data   = f_upbit.result()
print(f"[{time.time()-script_start:.1f}s] 외부 API 완료")


# ----------------------
# 5. 데이터프레임 조립
# ----------------------
def build_kr_df():
    rows = []
    for group_name, group_items in kr_groups:
        rows.append([group_name, "", "", "", ""])
        for name, code in group_items:
            d = cache.get(f"{code}.KS", {})
            rows.append([
                name,
                code,
                d.get("close") or "",
                d.get("change_pct", "-"),
                    d.get("dividend_yield") or "0.00%",
            ])
    return pd.DataFrame(rows, columns=["종목", "코드", "현재가", "변동률(%)", "배당률(%)"])

def build_us_df():
    rows = []
    for group_name, group_items in us_groups:
        rows.append([group_name, "", "", ""])
        for ticker in group_items:
            d = cache.get(ticker, {})
            rows.append([
                ticker,
                d.get("close") or "",
                d.get("change_pct", "-"),
                d.get("dividend_yield") or "0.00%",
            ])
    return pd.DataFrame(rows, columns=["종목", "현재가", "변동률(%)", "배당률(%)"])

def build_jp_df():
    rows = []
    for name, ticker in jp_stocks:
        d = cache.get(ticker, {})
        rows.append([
            name,
            ticker,
            d.get("close") or "",
            d.get("change_pct", "-"),
            d.get("dividend_yield") or "0.00%",
        ])
    return pd.DataFrame(rows, columns=["종목", "코드", "현재가", "변동률(%)", "배당률(%)"])

def build_other_df():
    # 기본 포맷 함수: 천 단위 콤마 + 소수점 2째 자리
    fmt2 = lambda p: f"{float(p):,.2f}" if p is not None else ""
    # 소수점 4자리 포맷
    fmt4 = lambda p: f"{float(p):,.4f}" if p is not None else ""
    # 소수점 제거 포맷
    fmt0 = lambda p: f"{int(float(p)):,}" if p is not None else ""
    
    rows = []
    # 환율 데이터
    for label, key in [("USD/KRW","KRW"),("USD/JPY","JPY"),("USD/CNY","CNY"),("USD/HKD","HKD")]:
        p = rate_data.get(key)
        rows.append([label, fmt2(p)])

    # 암호화폐 데이터 (USD)
    for label, key in [("USD/BTC","BTC"),("USD/ETH","ETH"),("USD/USDT","USDT"),("USD/NEXO","NEXO")]:
        p = crypto_data.get(key)
        if key == "NEXO":
            rows.append([label, fmt4(p)]) # B9: NEXO는 4자리
        else:
            rows.append([label, fmt2(p)])

    # 업비트 데이터 (KRW)
    for label, market in [("비트코인(업비트)","KRW-BTC"),("이더리움(업비트)","KRW-ETH"),("테더(업비트)","KRW-USDT")]:
        p = upbit_data.get(market)
        rows.append([label, fmt0(p)]) # B10~B12: 업비트는 소수점 제거

    return pd.DataFrame(rows, columns=["종목", "가격"])

df_kr    = build_kr_df()
df_us    = build_us_df()
df_jp    = build_jp_df()
df_other = build_other_df()

# ----------------------
# 6. 엑셀 저장 + 서식
# ----------------------
file_path = Path(r"C:\Users\MSJ\Documents\[상준 폴더]\[Investment]\주식 가격표(python)3_claude.xlsx")
file_path.parent.mkdir(parents=True, exist_ok=True)
update_date = datetime.now().strftime("%Y-%m-%d")
update_time = datetime.now().strftime("%H:%M:%S")

KR_GROUPS = {"<삼성증권 퇴직연금>","<Toss>","<NH농협 개인연금>","<미래에셋 개인연금>","<미래에셋 IRP>","<NH ISA>"}
US_GROUPS = {"<Creon>","<Meritz>","<Toss>"}

GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
BOLD_WHITE = Font(bold=True, color="FFFFFF", name="Arial")
BODY_FONT  = Font(name="Arial", size=10)
HEADER_FONT = Font(name="Arial", size=10, bold=True)

def style_sheet(ws, group_markers: set, n_cols: int):
    # 헤더 행 서식
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER
        cell.font = BOLD_WHITE
        cell.alignment = CENTER

    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        is_group = val in group_markers
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CENTER
            cell.font = BODY_FONT
            if is_group:
                cell.fill = GREEN
                cell.font = Font(name="Arial", size=10, bold=True)
            # 숫자 서식
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                if col == 3 and n_cols == 5:   # 한국주식 현재가 → 정수
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00" if isinstance(cell.value, float) else "#,##0"

    # 업데이트 일시
    ws["G1"] = update_date
    ws["H1"] = update_time
    for cell in (ws["G1"], ws["H1"]):
        cell.alignment = CENTER
        cell.font = HEADER_FONT

    # 열 너비 자동 조정
    for col_idx in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value))
             for r in range(1, ws.max_row + 1)
             if ws.cell(row=r, column=col_idx).value is not None),
            default=8,
        )
        ws.column_dimensions[col_letter].width = max_len + 4

print(f"[{time.time()-script_start:.1f}s] 엑셀 저장 중...")
try:
    # ExcelWriter를 사용하여 여러 데이터프레임을 하나의 엑셀 파일 내 여러 시트로 저장합니다.
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df_kr.to_excel(writer,    sheet_name="한국주식", index=False)
        df_us.to_excel(writer,    sheet_name="미국주식", index=False)
        df_jp.to_excel(writer,    sheet_name="일본주식", index=False)
        df_other.to_excel(writer, sheet_name="기타",    index=False)

        # 저장된 각 시트에 디자인(헤더 색상, 정렬 등)을 적용합니다.
        style_sheet(writer.sheets["한국주식"], KR_GROUPS, n_cols=5)
        style_sheet(writer.sheets["미국주식"], US_GROUPS, n_cols=4)
        style_sheet(writer.sheets["일본주식"], set(),     n_cols=5)
        style_sheet(writer.sheets["기타"],    set(),     n_cols=2)

    total = time.time() - script_start
    print(f"✅ 완료! 총 소요 시간: {total:.1f}초  →  {file_path}")

except PermissionError:
    # 만약 사용자가 엑셀 파일을 열어두고 있다면 쓰기 권한 오류가 발생합니다.
    print(f"\n❌ 오류: '{file_path.name}' 파일이 이미 열려 있어 저장할 수 없습니다. 파일을 닫고 다시 실행해 주세요.")
except Exception as e:
    # 기타 예외 상황에 대한 처리
    print(f"\n❌ 저장 중 예상치 못한 오류가 발생했습니다: {e}")